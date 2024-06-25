import os
import pandas as pd
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ensure the output directory exists
output_dir = "results/correlation/sorted"
os.makedirs(output_dir, exist_ok=True)

# Load the full correlation results CSV file
correlation_results_path = "results/correlation/attribute_correlations_full.csv"
correlation_results = pd.read_csv(correlation_results_path)

# Remove constant columns
data = pd.read_csv("CSVs/sample_train_100.csv")
data = data.loc[:, (data != data.iloc[0]).any()]
attributes = data.columns

# Normalize the correlation values
scaler = MinMaxScaler()
correlation_results[
    ["Norm_Pearson_Correlation", "Norm_Spearman_Correlation", "Norm_Mutual_Information"]
] = scaler.fit_transform(
    correlation_results[
        ["Pearson_Correlation", "Spearman_Correlation", "Mutual_Information"]
    ]
)


# Define a function to apply formatting
def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Resize columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Highlight the top 5 rows
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in range(2, 7):  # Highlight rows 2 to 6 (top 5)
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    wb.save(file_path)


# Prepare a dictionary to store sorted results for each attribute
sorted_results = {}

# Set to track processed pairs to avoid duplicates
processed_pairs = set()

# Sort the correlations for each attribute
for attr in attributes:
    # Filter the correlations where attr is either Attribute1 or Attribute2
    filtered_results = correlation_results[
        (correlation_results["Attribute1"] == attr)
        | (correlation_results["Attribute2"] == attr)
    ].copy()

    # Add 'attr2' column
    filtered_results["attr2"] = filtered_results.apply(
        lambda row: (
            row["Attribute2"] if row["Attribute1"] == attr else row["Attribute1"]
        ),
        axis=1,
    )

    # Keep only the necessary columns and reorder them
    filtered_results = filtered_results[
        [
            "attr2",
            "Pearson_Correlation",
            "Spearman_Correlation",
            "Mutual_Information",
            "Norm_Pearson_Correlation",
            "Norm_Spearman_Correlation",
            "Norm_Mutual_Information",
        ]
    ]
    filtered_results.insert(0, "Attribute1", attr)

    # Calculate Combined_Score using normalized values
    filtered_results["Combined_Score"] = (
        filtered_results["Norm_Pearson_Correlation"]
        + filtered_results["Norm_Spearman_Correlation"]
        + filtered_results["Norm_Mutual_Information"]
    ) / 3

    # Sort by Combined_Score
    sorted_filtered_results = filtered_results.sort_values(
        by="Combined_Score", ascending=False
    )

    # Remove duplicates by keeping track of processed pairs
    sorted_filtered_results = sorted_filtered_results[
        sorted_filtered_results.apply(
            lambda row: (row["Attribute1"], row["attr2"]) not in processed_pairs
            and not processed_pairs.add((row["Attribute1"], row["attr2"])),
            axis=1,
        )
    ]

    # Store the sorted results
    sorted_results[attr] = sorted_filtered_results

    # Save the sorted results to an Excel file for each attribute
    output_file_path = f"{output_dir}/sorted_{attr}_correlations.xlsx"
    sorted_filtered_results.to_excel(output_file_path, index=False)

    # Apply formatting
    format_excel(output_file_path)

    print(f"Sorted correlation results for {attr} saved to {output_file_path}")

# Optionally: Save all sorted results into a single Excel file
combined_sorted_results = pd.concat(sorted_results.values(), keys=sorted_results.keys())
combined_output_file_path = f"{output_dir}/sorted_attribute_correlations_full.xlsx"
combined_sorted_results.to_excel(combined_output_file_path, index=False)
format_excel(combined_output_file_path)
print(f"Combined sorted correlation results saved to {combined_output_file_path}")
